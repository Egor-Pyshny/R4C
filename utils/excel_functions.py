import os
from datetime import datetime
from datetime import timedelta
from typing import Dict
from typing import Tuple

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from robots.models import Robot


def get_data() -> Tuple[Dict[str, int], Dict[str, Dict[str, int]]]:
    current_date = datetime.now().replace(second=0, microsecond=0)
    one_week_ago = current_date - timedelta(days=7)
    general_report_data: Dict[str, int] = {}
    detailed_report_data: Dict[str, Dict[str, int]] = {}
    for robot in Robot.fetch_data():
        created = robot.created
        version = robot.version
        model = robot.model
        if created >= one_week_ago:
            if model not in detailed_report_data.keys():
                detailed_report_data[model] = {}
                general_report_data[model] = 1
            else:
                general_report_data[model] += 1
            versions = detailed_report_data[model]
            if version in versions.keys():
                detailed_report_data[model][version] += 1
            else:
                detailed_report_data[model][version] = 1
        else:
            break
    return general_report_data, detailed_report_data


def create_general_report(general_report_data: Dict[str, int]) -> None:
    workbook: Workbook = Workbook()
    worksheet: Worksheet = workbook["Sheet"]
    worksheet.title = "General report"
    worksheet.append(["Model", "Amount"])
    for node in general_report_data.keys():
        worksheet.append([node, general_report_data[node]])
    workbook.save(str(os.environ.get("PATH_FOR_ROBOT_REPORT")))


def create_detailed_report(
    detailed_report_data: Dict[str, Dict[str, int]]
) -> None:
    workbook: Workbook = load_workbook(
        str(os.environ.get("PATH_FOR_ROBOT_REPORT"))
    )
    for model in detailed_report_data.keys():
        worksheet: Worksheet = workbook.create_sheet(model + " report")
        worksheet.append(["Model", "Version", "Amount"])
        versions: Dict[str, int] = detailed_report_data[model]
        for version in versions.keys():
            worksheet.append([model, version, versions[version]])
    workbook.save(str(os.environ.get("PATH_FOR_ROBOT_REPORT")))


def create_report() -> None:
    general_report_data, detailed_report_data = get_data()
    create_general_report(general_report_data)
    create_detailed_report(detailed_report_data)
